"""Export all UI + zone strings (KO + EN) into a single xlsx workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

OUT = Path("docs/i18n_strings.xlsx")
OUT.parent.mkdir(exist_ok=True)

# ── UI strings ────────────────────────────────────────────────────────────────
ui_rows = [
    # category, key, location, ko, en
    ("Splash", "splashLabel",        "Splash · 상단 라벨",                     "강릉 이머시브 아트 쇼", "GANGNEUNG IMMERSIVE ART SHOW"),
    ("Splash", "splashSloganLine1",  "Splash · 슬로건 (1줄)",                  "다섯개의 달이 피어오르는 밤,", "On a night when five moons rise,"),
    ("Splash", "splashSloganLine2",  "Splash · 슬로건 (2줄)",                  "고대 하슬라의 판타지가 펼쳐집니다.", "the fantasy of ancient Hasla unfolds."),
    ("Splash", "splashHint",         "Splash · 탭 안내",                       "TAP TO AWAKEN", "TAP TO AWAKEN"),
    ("Splash", "splashCta",          "Splash · CTA 버튼",                      "숲으로 들어가기", "Enter the Forest"),
    ("Splash", "copyrightSuffix",    "Splash · 푸터 표기",                      "강릉 경포 환상의 호수", "Gyeongpo Fantastic Lake, Gangneung"),

    ("Header", "navHome",            "맵 페이지 헤더 좌측",                     "← HOME", "← HOME"),
    ("Header", "navMap",             "Zone 상세 헤더 좌측",                    "← MAP", "← MAP"),

    ("Map",    "toggleMap",          "지도/목록 토글 — 지도",                   "지도보기", "Map"),
    ("Map",    "toggleList",         "지도/목록 토글 — 목록",                   "목록보기", "List"),
    ("Map",    "mapCaption",         "맵 하단 안내 문구",                       "빛나는 영역을 탭하면 각 ZONE의 안내로 이동합니다.", "Tap a glowing zone to view its details."),
    ("Map",    "viewDetail",         "리스트 카드 — 자세히 보기 링크",          "자세히 보기 →", "View details →"),

    ("Zone",   "sectionDirection",   "Zone 상세 — 연출 방향 헤딩",              "연출 방향", "Direction"),
    ("Zone",   "sectionMedia",       "Zone 상세 — 연출 매체 헤딩",              "연출 매체", "Production Media"),
    ("Zone",   "navBackToMap",       "Zone 상세 — 마지막 zone에서 맵으로",      "맵으로 돌아가기", "Back to map"),
    ("Zone",   "navFinish",          "Zone 상세 — 마지막 zone 라벨",            "FINISH →", "FINISH →"),
    ("Zone",   "navFirst",           "Zone 상세 — 첫 zone에서 홈으로 라벨",      "처음으로", "Home"),
    ("Zone",   "zoneTotalSeparator", "Zone 상세 — 페이지 표시 구분자",          "/", "/"),

    ("Audio",  "audioTitleDefault",  "오디오 플레이어 — 기본 제목",             "오디오 도슨트", "Audio Docent"),
    ("Audio",  "audioTapToPlay",     "오디오 플레이어 — 재생 안내",             "화면을 탭하면 재생됩니다.", "Tap the screen to play."),
    ("Audio",  "audioError",         "오디오 플레이어 — 에러",                  "오디오 파일이 아직 준비되지 않았습니다.", "Audio is not yet available."),
    ("Audio",  "ariaPlay",           "오디오 플레이어 — aria-label 재생",       "재생", "Play"),
    ("Audio",  "ariaPause",          "오디오 플레이어 — aria-label 일시정지",    "일시정지", "Pause"),
    ("Audio",  "ariaSeek",           "오디오 플레이어 — aria-label 시크",       "재생 위치", "Seek position"),

    ("Element", "elementWater",      "오행 — 수",                              "수(水)", "Water"),
    ("Element", "elementWood",       "오행 — 목",                              "목(木)", "Wood"),
    ("Element", "elementEarth",      "오행 — 토",                              "토(土)", "Earth"),
    ("Element", "elementFire",       "오행 — 화",                              "화(火)", "Fire"),
    ("Element", "elementMetal",      "오행 — 금",                              "금(金)", "Metal"),

    ("Meta",    "metaTitle",         "행사 단축명",                            "경포 환상의 호수", "Gyeongpo Fantastic Lake"),
    ("Meta",    "metaSubtitle",      "행사 부제",                              "고대 하슬라의 밤", "Night of Ancient Hasla"),
    ("Meta",    "metaFullTitle",     "행사 풀네임",                            "경포 환상의 호수 — 고대 하슬라의 밤", "Gyeongpo Fantastic Lake — Night of Ancient Hasla"),
    ("Meta",    "slogan",            "메타 슬로건",                            "다섯개의 달이 피어오르는 밤, 고대 하슬라의 판타지가 펼쳐집니다", "On a night when five moons rise, the fantasy of ancient Hasla unfolds"),
    ("Meta",    "concept",           "행사 컨셉 한 줄",                         "다섯 개의 달이 뜬 밤, 고대 하슬라의 모습이 재현된다.", "When five moons rise, the world of ancient Hasla returns."),

    ("Misc",    "langToggleLabel",   "언어 토글 라벨",                          "LANGUAGE", "LANGUAGE"),
]

# ── Zone strings ──────────────────────────────────────────────────────────────
# (zone_id, field, ko, en)
zones_ko_en = {
    "zone1": {
        "title":       ("하슬라 포털",                                                              "Hasla Portal"),
        "subtitle":    ("달빛이 열어주는 통로",                                                       "A passage opened by moonlight"),
        "tagline":     ("달 모양의 입구를 통해 신비로운 세상과 통하는 하슬라 포털",                        "A moon-shaped gateway leading into a mystical world."),
        "story":       ("고대 하슬라의 숲으로 향하는 게이트 중 하나, 달빛이 열어주는 달을 닮은 통로",          "One of the gateways into the forest of ancient Hasla — a moon-shaped passage opened by moonlight."),
        "description": ("송림 사이에 나타난 원형의 포털은 하슬라 세계로 들어가는 관문이자 시작점입니다. 하슬라의 신화와 이야기는 이곳에서 열리며, 포털을 지나면 판타지의 세계가 펼쳐집니다.",
                        "Set among the pines, this circular portal is the threshold into the world of Hasla — the place where its myths begin. Step through, and a realm of fantasy unfolds before you."),
        "direction.0": ("달의 형상을 닮은 문을 지나며, 관람객은 현실의 숲에서 신화 속 하슬라의 세계로 진입한다.",
                        "Visitors pass through a moon-shaped gate, crossing from the real forest into the mythic world of Hasla."),
        "direction.1": ("빛과 안개가 어우러진 링 형태의 포털을 통해, 관람객이 고대 하슬라의 세계로 진입하는 듯한 체험을 제공한다.",
                        "A ring-shaped portal of light and fog evokes the sensation of stepping into ancient Hasla."),
        "media.0":     ("라인 조명",       "Line lighting"),
        "media.1":     ("원형 입구 조형물",  "Circular entrance structure"),
        "media.2":     ("포그머신",         "Fog machine"),
    },
    "zone2": {
        "title":       ("그루터기의 숨결",                                                          "Breath of the Stumps"),
        "subtitle":    ("깨어나 빛으로 맥박하는 그루터기",                                            "Stumps awakening, pulsing with light"),
        "tagline":     ("그루터기가 관람객의 두드림에 화답하며, 빛과 화음으로 숲의 생명을 일깨운다",         "Each stump answers a visitor's touch — light and harmony awakening the forest's life."),
        "story":       ("당시 하슬라는 달빛 에너지를 받아 모든 동식물이 자유롭게 노래하던 시절이었다. 그루터기에 손을 얹으면 빛으로 맥박하며 깨어고, 각자의 하모니를 쌓아 노래하며 숲의 복귀를 축하한다.",
                        "In ancient Hasla, drawn by moonlight energy, every living thing once sang freely. Lay a hand on a stump and it pulses awake with light, layering harmonies that celebrate the forest's return."),
        "description": ("달빛이 스며들며 하슬라의 숲이 다시 숨을 쉽니다. 고대의 기억이 남은 그루터기는 노래로 화답하고, 관람객이 그루터기를 두드리면 빛과 사운드가 파동처럼 확산되며 숲의 맥박이 살아나는 연출이 구현됩니다. 관람객 참여를 통해 생명의 기운이 깨어나는 경험을 제공합니다.",
                        "Moonlight seeps in, and the Hasla forest breathes again. Stumps that hold ancient memory respond with song; tap them, and waves of light and sound ripple outward — the forest's pulse coming alive. Through participation, visitors experience life itself awakening."),
        "direction.0": ("각 오브제(그루터기·연못)가 반응형 인터랙션으로 구현되어 관람객이 ‘깨어남’의 순간을 직접 체험.",
                        "Each object (stumps, ponds) responds interactively, letting visitors experience the moment of awakening directly."),
        "direction.1": ("체험자가 두드리는 타이밍에 맞춰 조명·음향이 연동되며 숲이 맥박하듯 주변으로 빛이 퍼져 나간다.",
                        "Lighting and sound sync to each tap; light radiates outward like the forest's pulse."),
        "media.0":     ("모형 그루터기",      "Stump models"),
        "media.1":     ("라인 조명",         "Line lighting"),
        "media.2":     ("인터랙션 센서",      "Interaction sensors"),
        "media.3":     ("스피커",            "Speakers"),
    },
    "zone3": {
        "title":       ("거울 연못", "Mirror Pond"),
        "subtitle":    ("달빛을 받아 재현된 하슬라의 연못", "The pond of Hasla, restored by moonlight"),
        "tagline":     ("달빛을 받은 정원의 연못 위에 하슬라의 생명이 비치고, 관람객이 다가서면 움직임에 반응한다",
                        "On the moonlit pond, the life of Hasla appears — and stirs as you approach."),
        "story":       ("달빛 아래 신비한 생명들(빛의 물고기)이 모여드는 비밀의 샘, 다가서면 과거 하슬라 시절의 연못이 재현된다.",
                        "A secret spring beneath the moon, where mystical lives — fish of light — gather. Step closer, and the pond of old Hasla is restored before your eyes."),
        "description": ("거울면 위로 달빛과 생명체의 영상이 투사되며, 관람객의 접근에 따라 물결·빛·생명체가 반응하는 인터랙션이 연출됩니다. 가까이 다가서면 도망가는 빛의 물고기를 따라 달빛에 잠긴 연못의 시간을 거닐어 보세요.",
                        "Moonlight and creatures are projected onto the mirrored surface; ripples, lights, and beings respond as you draw near. Follow the fleeing fish of light and walk through time on a pond bathed in moonlight."),
        "direction.0": ("스마트 미러 위에 헤엄치는 하슬라의 물고기 영상.",  "Smart mirrors display Hasla's fish swimming across their surface."),
        "direction.1": ("관람객의 접근에 따라 도망가는 물고기 — 인터랙션 센서로 연동.",  "Interaction sensors send the fish fleeing as visitors approach."),
        "media.0":     ("스마트 미러",      "Smart mirror"),
        "media.1":     ("거울",             "Mirror panels"),
        "media.2":     ("인터랙션 센서",     "Interaction sensors"),
        "media.3":     ("스피커",           "Speakers"),
    },
    "zone4": {
        "title":       ("다섯 개의 달", "Five Moons"),
        "subtitle":    ("대왕 그루터기 — 하슬라 숲의 심장", "The Great Stump — heart of the Hasla forest"),
        "tagline":     ("밤하늘로부터 쏟아지는 다섯개의 달빛을 받아내는, 세상에서 가장 큰 소나무 그루터기",
                        "The world's greatest pine stump, catching five streams of moonlight from the night sky."),
        "story":       ("다섯 달빛을 집중적으로 받아 세상에서 가장 크게 자라난, 하슬라 숲의 심장",
                        "The heart of the Hasla forest — grown to colossal size by the concentrated light of five moons."),
        "description": ("하슬라 시절 가장 큰 소나무가 있던 자리. 거대한 그루터기에서 다섯 줄기의 빛이 수직으로 솟아오르며 다섯 달을 상징합니다. 서치라이트 조명 연출이 잠들었던 하슬라의 부활을 알리고, 신비한 다섯개의 달빛은 하슬라 전역으로 퍼져 나갑니다 — 이 강력한 빛은 하슬라 생태계를 구성하는 모든 에너지의 근원입니다.",
                        "Where the greatest pine of Hasla once stood. From the giant stump, five vertical beams rise to signify the five moons. Searchlights announce the rebirth of slumbering Hasla as the mystical moonlight radiates across the forest — the primal energy that sustains all of Hasla's ecology."),
        "direction.0": ("서치라이트로 하슬라 대표 컨셉인 다섯 개의 달을 강조.",  "Searchlights highlight Hasla's signature concept — the Five Moons."),
        "direction.1": ("높게 뜬 다섯 개의 달 연출을 통해 하슬라 세계관의 상징적 의미 전달.",  "Five moons rise high, conveying the symbolic core of the Hasla world."),
        "media.0":     ("서치라이트",            "Searchlights"),
        "media.1":     ("대형 그루터기 조형물",   "Giant stump structure"),
        "media.2":     ("사운드 시스템",         "Sound system"),
    },
    "zone5": {
        "title":       ("달의 초상", "Portrait of the Moon"),
        "subtitle":    ("미러 벽체 — 거울에 비친 달", "Mirror walls — the moon caught in reflection"),
        "tagline":     ("하늘에 뜬 달을 거울 속에서 마주하는 신비로운 공간",  "A mystical chamber where you meet the moon through its reflection."),
        "story":       ("달빛은 하슬라 숲의 근원이었기에, 사람들은 이 거울에 비친 달의 초상을 통해서만 달과 조우할 수 있었다. (거울에 비친 달에게 소원을 빌거나 등..)",
                        "Because moonlight was the source of the Hasla forest, people could meet the moon only through its reflection in this mirror — and many made their wishes there."),
        "description": ("조명쇼 구간으로 넘어가기 전 시야를 전환하는 전이 공간. 지그재그로 배치된 거울 벽 사이를 걸으며, 관람객은 달빛에 비친 자신의 모습과 달의 잔영을 마주합니다. 반사면에 비친 달과 관람객의 모습이 겹쳐지며 달빛과 내가 하나가 되는 경험을 제공하고, 야간 포토존으로도 활용됩니다.",
                        "A transitional space before the light show. Walking between zigzag mirror walls, visitors meet themselves bathed in moonlight alongside the moon's afterimage — a moment when self and moon become one. It also doubles as a nightscape photo spot."),
        "direction.0": ("지그재그 배치된 미러월로 시야를 차폐하고 동선을 유도.",  "Zigzag mirror walls block the view and guide the path."),
        "direction.1": ("거울에 비친 달빛과 반사로 관람객이 달을 마주하는 경험 구현.",  "Reflected moonlight lets visitors come face to face with the moon."),
        "media.0":     ("거울 벽체",  "Mirror walls"),
    },
    "zone6": {
        "title":       ("인피니티 포레스트", "Infinity Forest"),
        "subtitle":    ("달빛의 창 — 빛의 축제", "Window of moonlight — a festival of light"),
        "tagline":     ("7M의 LED WALL로 연출하는 몰입형 콘텐츠 — 다섯 개의 달이 차례로 떠오르는 문라이트 쇼",
                        "An immersive 7m LED wall presenting the Moonlight Show — five moons rising one by one."),
        "story":       ("달빛에너지와 숲이 음악과 빛으로 어우러져, 고대 하슬라 숲의 신비로운 세계를 재현하는 체험의 장. 다섯 달의 세계를 비추는 거대한 포털, 하늘에 뜬 달들을 엿보는 신비로운 창.",
                        "An experiential stage where moonlight, forest, music, and light merge to recreate the mystical world of ancient Hasla — a vast portal onto the realm of five moons, a window into the sky's hidden lights."),
        "description": ("송림 사이에 설치된 약 7.4m 규모의 LED WALL로 끝없이 확장되는 숲을 구현한 공간. 빛과 음향이 어우러지며 관람객은 신비로운 세계 속을 거니는 듯한 몰입감을 경험합니다. 다섯 달이 차례로 떠오르고 마침내 하나로 모이는 ‘하슬라 문라이트 쇼'가 펼쳐집니다.",
                        "A roughly 7.4m LED wall set among the pines transforms the forest into an endless expanse. Light and sound combine in an immersive walk through a mystical world. Five moons rise one by one, finally uniting in the climactic Hasla Moonlight Show."),
        "direction.0": ("EVERGREEN MOON(목) — 푸른 달빛이 숲을 적시며 사슴과 산맥이 모습을 드러낸다.",  "EVERGREEN MOON (Wood) — Blue moonlight bathes the forest as deer and mountain ranges appear."),
        "direction.1": ("SOLAR MOON(화) — 태양과 달이 교차하며 붉은 불꽃이 피어오른다.",  "SOLAR MOON (Fire) — Sun and moon cross paths and red flames rise."),
        "direction.2": ("AMBER MOON(토) — 불길의 재가 황토가 되어 새로운 씨앗이 움튼다.",  "AMBER MOON (Earth) — Embers turn to ochre soil and new seeds sprout."),
        "direction.3": ("STARLIGHT MOON(금) — 흙 속에서 태어난 보석이 별이 되어 밤하늘을 물들인다.",  "STARLIGHT MOON (Metal) — Gems born in the earth ascend to the sky as stars."),
        "direction.4": ("CLOUD MOON(수) — 별빛 사이 구름이 몰려와 빗방울이 숲을 적신다.",  "CLOUD MOON (Water) — Clouds sweep through starlight and rain falls upon the forest."),
        "direction.5": ("THE NEW MOON(合) — 다섯 빛이 하나의 신월로 모여 천문도의 무늬를 새긴다.",  "THE NEW MOON (Union) — Five lights merge into one new moon, etching a star-chart pattern across the sky."),
        "media.0":     ("대형 LED WALL (7.4m)",   "Large LED wall (7.4m)"),
        "media.1":     ("LED 바",                 "LED bars"),
        "media.2":     ("조명기 일체",             "Full lighting rig"),
        "media.3":     ("포그머신",               "Fog machines"),
        "media.4":     ("사운드 시스템",           "Sound system"),
    },
    "zone7": {
        "title":       ("달의 잔상",  "Afterimage of the Moon"),
        "subtitle":    ("원형 데크 위에서 펼쳐지는 다섯 달의 세계", "The world of five moons, unfolding on a circular deck"),
        "tagline":     ("원형 데크 위에 떠오른 찬란한 달빛, 발 밑에서 펼쳐지는 또 하나의 신비로운 세계",  "Brilliant moonlight on a circular deck — another mystical world unfolding beneath your feet."),
        "story":       ("관람객이 하슬라의 달빛과 그 세계를 걸으며 숲의 기억과 교감하는 인터랙티브 원형 공간",  "An interactive circular space where visitors walk through Hasla's moonlight and commune with the memory of the forest."),
        "description": ("원형 데크에 프로젝션 맵핑을 적용한 몰입형 인터랙션 공간. 중앙의 달은 조명쇼 각 테마와 연동되어 변화하고, 바닥에는 달의 세계가 투사됩니다. 관람객의 움직임에 반응해 파도와 균열, 빛의 파편이 피어나며 — 마치 사라진 달의 기억 속을 걷는 듯한 잔상의 세계를 마주하게 됩니다.",
                        "An immersive interactive space with projection mapping on a circular deck. The central moon shifts in tune with each theme of the light show, while the moon's world is cast across the floor. Movement triggers waves, fissures, and light fragments — as if walking through the lingering memory of a vanished moon."),
        "direction.0": ("원형 데크 중앙에 떠오른 달과 주변 세계가 시시각각 변주.",  "The central moon and the world around it shift moment by moment."),
        "direction.1": ("관람객의 발걸음에 맞춰 빛의 파편과 파동이 피어나는 몰입형 연출.",  "Light fragments and ripples bloom in response to footsteps."),
        "direction.2": ("달을 중심으로 풍경이 은은하게 흐르고, 잔상처럼 부드러운 이펙트가 겹쳐져 여운을 남기는 서정적 연출.",  "A lyrical scene flows around the moon, with soft afterimage effects layering into a quiet resonance."),
        "media.0":     ("빔 프로젝터 2EA",          "Beam projector × 2"),
        "media.1":     ("프로젝션 맵핑 시스템",      "Projection mapping system"),
        "media.2":     ("위치 감지 센서",            "Position sensors"),
        "media.3":     ("사운드 시스템",            "Sound system"),
    },
    "zone8": {
        "title":       ("빛의 파동",  "Wave of Light"),
        "subtitle":    ("달의 에너지가 숲길을 따라 흐르는 순간",  "The moment moonlight flows along the forest path"),
        "tagline":     ("바닥 위에 흘러내리는 빛의 파동",  "Waves of light pouring across the ground."),
        "story":       ("달빛의 에너지가 숲 바닥을 타고 물결치듯 흘러, 고대 하슬라의 숲을 깨어나게 한다.",  "Moonlight energy flows across the forest floor in waves, awakening the ancient forest of Hasla."),
        "description": ("송림 바닥에 달의 에너지가 흐르는 듯한 분위기를 조성하는 고보 조명 연출 공간. 바닥에 고보 패턴을 투사해 신비로운 달빛의 파동이 송림 바닥을 따라 흘러갑니다. 한 걸음 내딛을 때마다 달의 파동이 바닥에서 울려 퍼지고 — 관람객은 빛과 함께 걷는 특별한 경험을 하게 됩니다.",
                        "Gobo lighting makes the moon's energy seem to flow along the pine forest floor. Mystical waves of moonlight ripple across the ground; with every step, the moon's pulse echoes — and visitors walk in step with the light itself."),
        "direction.0": ("바닥 고보 조명을 통해 달빛이 물결처럼 번지며 숲길 전체에 파동을 형성.",  "Gobo lighting spreads moonlight in ripples that form waves across the path."),
        "direction.1": ("달빛의 파동이 바닥을 따라 이어지며 관람객의 이동 경로를 자연스럽게 안내하는 동선 유도형 연출.",  "The waves of moonlight guide visitors gently along the route."),
        "media.0":     ("고보 라이팅",     "Gobo lighting"),
        "media.1":     ("사운드 시스템",   "Sound system"),
    },
}

# ── Workbook ─────────────────────────────────────────────────────────────────
wb = Workbook()

font_default = Font(name='Arial', size=10)
font_bold = Font(name='Arial', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', start_color='2353A3')
zebra_fill = PatternFill('solid', start_color='F4F6FA')
border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def style_header_row(ws, last_col_letter):
    for col in 'ABCDEFGH'[: ord(last_col_letter) - ord('A') + 1]:
        c = ws[f'{col}1']
        c.font = font_bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[1].height = 28

def style_body(ws, last_col_letter):
    last_col_idx = ord(last_col_letter) - ord('A') + 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=last_col_idx), start=2):
        is_zebra = row_idx % 2 == 0
        for c in row:
            c.font = font_default
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = border
            if is_zebra:
                c.fill = zebra_fill

# Sheet 1: UI strings
ws_ui = wb.active
ws_ui.title = 'UI Strings'
ws_ui.append(['Category', 'Key', 'Location', 'Korean (KO)', 'English (EN)'])
for row in ui_rows:
    ws_ui.append(list(row))
ws_ui.column_dimensions['A'].width = 14
ws_ui.column_dimensions['B'].width = 22
ws_ui.column_dimensions['C'].width = 32
ws_ui.column_dimensions['D'].width = 50
ws_ui.column_dimensions['E'].width = 60
ws_ui.freeze_panes = 'A2'
style_header_row(ws_ui, 'E')
style_body(ws_ui, 'E')

# Sheet 2: Zone strings
ws_z = wb.create_sheet('Zone Content')
ws_z.append(['Zone', 'Field', 'Korean (KO)', 'English (EN)'])
for zid, fields in zones_ko_en.items():
    for fname, (ko, en) in fields.items():
        ws_z.append([zid, fname, ko, en])
ws_z.column_dimensions['A'].width = 10
ws_z.column_dimensions['B'].width = 18
ws_z.column_dimensions['C'].width = 70
ws_z.column_dimensions['D'].width = 80
ws_z.freeze_panes = 'A2'
style_header_row(ws_z, 'D')
style_body(ws_z, 'D')

# Sheet 3: Summary cover
ws_sum = wb.create_sheet('Overview', 0)
ws_sum['A1'] = '경포 환상의 호수 — 모바일 리플렛 i18n 문자열'
ws_sum['A1'].font = Font(name='Arial', size=16, bold=True, color='2353A3')
ws_sum.merge_cells('A1:D1')
ws_sum['A2'] = 'Gyeongpo Fantastic Lake — Mobile Leaflet i18n Strings'
ws_sum['A2'].font = Font(name='Arial', size=11, italic=True, color='666666')
ws_sum.merge_cells('A2:D2')

ws_sum['A4'] = 'Sheet'
ws_sum['B4'] = 'Description'
ws_sum['C4'] = 'Rows'
for col in 'ABC':
    ws_sum[f'{col}4'].font = font_bold
    ws_sum[f'{col}4'].fill = header_fill
    ws_sum[f'{col}4'].border = border
    ws_sum[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
ws_sum.row_dimensions[4].height = 24

# Use formulas for row counts
ws_sum.append(['UI Strings',  'Splash · 헤더 · 맵 · Zone 상세 · 오디오 · 오행 · 메타 · 기타 (UI 텍스트)', f'=COUNTA(\'UI Strings\'!A:A)-1'])
ws_sum.append(['Zone Content','각 Zone(8개)의 title/subtitle/tagline/story/description/direction[]/media[]', f'=COUNTA(\'Zone Content\'!A:A)-1'])

for r in (5, 6):
    for col in 'ABC':
        c = ws_sum[f'{col}{r}']
        c.font = font_default
        c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=True)

ws_sum.column_dimensions['A'].width = 18
ws_sum.column_dimensions['B'].width = 70
ws_sum.column_dimensions['C'].width = 12

ws_sum['A8'] = '※ 코드 적용 위치'
ws_sum['A8'].font = Font(name='Arial', size=10, bold=True)
ws_sum['A9'] = 'src/i18n/ui.ts        — UI Strings 시트의 Key/KO/EN'
ws_sum['A10'] = 'src/i18n/zones.ts    — Zone Content 시트의 zone × field × KO/EN'
ws_sum['A11'] = 'src/data/zones.json  — Zone Content의 KO 마스터 (참고용)'

wb.save(OUT)
print(f'wrote {OUT}')
